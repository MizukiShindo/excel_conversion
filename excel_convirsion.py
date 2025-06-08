import pandas as pd
import tkinter as tk #GUI構築ライブラリ
from tkinter import filedialog, messagebox  #ファイル選択ダイアログ
import openpyxl as px #pythonで.xlsxを開く追加ライブラリ
import xlrd as pxl #pythonで.xlsを開く追加ライブラリ
import os

#グローバル変数としてワークブックを保持する
wb = None
df = None

def read_excel_file(filepath):
    if filepath.endswith(".xls"):
        df = pd.read_excel(filepath, engine="xlrd")
    else:
        df = pd.read_excel(filepath, engine="openpyxl")

    return df

def detect_excel_format(filepath):
    with open(filepath, 'rb') as f:
        header = f.read(8)
    
    if header.startswith(b'\xD0\xCF\x11\xE0'):  # .xls (OLE Compound File Binary)
        return 'xls'
    elif header.startswith(b'PK'):  # .xlsx (ZIP-based)
        return 'xlsx'
    else:
        return 'unknown'
    
def attempt_repair_xls(filepath):
    try:
        df = pd.read_excel(filepath, engine="xlrd")
        save_path = filepath.replace(".xls", "_converted.xlsx")
        df.to_excel(save_path, index=False)
        messagebox.showinfo("修復成功", f"新しいExcelファイルとして保存しました:\n{save_path}")
        return save_path
    except Exception as e:
        messagebox.showerror("修復失敗", f"ファイル修復に失敗しました:\n{e}")
        return None

def open_file():
    global wb, df
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])

    if file_path:
        filename = os.path.basename(file_path)
        file_label.config(text=f"{filename}", fg ="blue", cursor="hand2")
        file_label.bind("<Button-1>", lambda e: os.startfile(file_path))
        ToolTip(file_label, "クリックでファイルを開く")

        if file_path.endswith(".xls"):
            try:
                df = pd.read_excel(file_path, engine="xlrd")
                wb = None
                analyze_columns_xls(df)
            except Exception:
                actual_format = detect_excel_format(file_path)
                if actual_format == 'xlsx':
                    repaired = attempt_repair_xls(file_path)
                    if repaired:
                        wb = px.load_workbook(repaired, data_only=True)
                        analyze_columns_xlsx(wb)
                else:
                    messagebox.showerror("エラー", "xlsファイルが破損しているか、形式が異なっています。")
        else:
            try:
                wb = px.load_workbook(file_path, data_only=True)
                analyze_columns_xlsx(file_path)
            except Exception as e:
                messagebox.showerror("エラー", f"ファイルを読み込めませんでした\n{e}")

    else:
        file_label.config(text="ファイル未選択")
        columns_text.set("")
        return

def analyze_columns_xls(df):
    result_lines = []

    for col in df.columns:
        count = df[col].apply(lambda x: isinstance(x, (int, float))).sum()
        result_lines.append(f"{col}: {count}件の数値データ")

    columns_text.set("\n".join(result_lines))

def analyze_columns_xlsx(filepath):
    try:
        if filepath.endswith(".xls"):
            df = pd.read_excel(filepath, engine="xlrd", header=None)
        else:
            wb = px.load_workbook(filepath, data_only=True)
            sheet = wb.active
            max_col = sheet.max_column
            max_row = sheet.max_row

            result_lines = []

            for col_idx in range(1, max_col +1):
                part1 = sheet.cell(row=1, column=col_idx).value or ""
                part2 = sheet.cell(row=1, column=col_idx).value or ""
                col_name = f"{part1}/{part2}".strip(" /")

                count = 0
                for row in range (2, max_row +1):
                    cell = sheet.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        count += 1

                result_lines.append(f"{col_name}: {count}件の数値データ")

            columns_text.set("\n".join(result_lines))
            return
    
        # xlsの場合（pandas使用）
        result_lines = []
        for col in df.columns:
            count = pd.to_numeric(df[col], errors='coerce').notna().sum()
            result_lines.append(f"{col_name}: {count}件の数値データ")
        
        columns_text.set("\n".join(result_lines))

    except Exception as e:
        messagebox.showerror("エラー", f"ファイルを読み込めませんでした\n{e}")

def save_file():
    global wb, df
    if wb is None and df is None:
        messagebox.showwarning("エラー", "まずExcelファイル(*.xlsxまたは*.xls)を開いてくださいヽ(`Д´)ﾉ")
        return

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="名前を付けて保存"
    )

    if not save_path:
        return  # キャンセルされた場合は何もしない

    try:
        if save_path.endswith(".xls"):
            messagebox.showwarning("保存形式エラー", "xls形式には保存できません。xlsx形式で保存してください。")
            return

        if df is not None:
            df.to_excel(save_path, index=False)
        else:
            wb.save(save_path)

        messagebox.showinfo("保存完了", f"Excelファイルを保存しました:\n{save_path}")
        print(f"ファイルを保存しました: {save_path}")

    except Exception as e:
        messagebox.showerror("保存失敗", f"ファイルの保存に失敗しました:\n{e}")
    
def quit_app():
    tk.Tk().withdraw()
    res = messagebox.askokcancel('アプリ終了', 'アプリを終了しますか？')

    if res is True:
        root.quit()

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        x, y, _, _ = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 20
        y = y + self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # 枠なし
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, background="#ffffe0", relief="solid", borderwidth=1, font=("Helvetica", 9))
        label.pack()

    def hide_tip(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

def main():
    global root, file_label, columns_text

    # GUIウィンドウ作成
    root = tk.Tk()
    root.title("Excelデータ変換")

    # ウィンドウの幅と高さ
    window_width = 600
    window_height = 400

    # 画面サイズの取得
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # 中央座標の計算
    x = int((screen_width - window_width) / 2)
    y = int((screen_height - window_height) / 2)

    # ウィンドウの位置を設定
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")

    #概要を表示する
    description = tk.Label(root, text="Excelファイルを読み込み、単位変換とグラフ描画を行います(´・ω・｀)", font=("Helvetica", 9))
    description.pack(padx = 10, pady =(10, 5))

    #「ファイルを開く」ボタン
    open_button = tk.Button(root, text="Excelファイルを開く", width=20, height=3, command=open_file)
    open_button.pack(padx=10, pady=10)

    #ファイル名表示用ラベル
    file_label = tk.Label(root, text="ファイル未選択", fg="blue", font=("Helvetica", 10))
    file_label.pack(padx=10, pady=10)

    columns_text = tk.StringVar()
    tk.Label(root, textvariable=columns_text, justify="left", wraplength=550, anchor="w").pack(pady=10)

    #「名前をつけて保存」ボタン
    save_button = tk.Button(root, text="保存先を指定して保存", width=20, height=3, command=save_file)
    save_button.pack(padx=20, pady=10)

    #「アプリを終了する」ボタン
    quit_button = tk.Button(text = "アプリを終了する", width=20, height=3, command = quit_app)
    quit_button.pack(padx=30, pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()