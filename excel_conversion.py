import pandas as pd
import tkinter as tk #GUI構築ライブラリ
from tkinter import filedialog, messagebox  #ファイル選択ダイアログ
import openpyxl as px #pythonで.xlsxを開く追加ライブラリ
import xlrd as pxl #pythonで.xlsを開く追加ライブラリ
import os
import subprocess
import sys

#グローバル変数としてワークブックを保持する
wb = None
df = None
file_path = None

def TF():#テスト用 TESTING FUNCTION
    messagebox.showinfo("TESTING FUNCTION", "THIS IS TESTING FUNCTION")


def show_splash_and_launch_main(): #起動画面の表示
    splash = tk.Tk()
    splash.title("wait...")

    # ウィンドウのサイズと位置
    splash_width = 400
    splash_height = 200
    screen_width = splash.winfo_screenwidth()
    screen_height = splash.winfo_screenheight()
    x = int((screen_width - splash_width) / 2)
    y = int((screen_height - splash_height) / 2)
    splash.geometry(f"{splash_width}x{splash_height}+{x}+{y}")
    splash.overrideredirect(True)  # 枠やボタン非表示にする

    # ラベルにタイトル表示
    label = tk.Label(splash, text="Excel Data Converter\nwait...", font=("Helvetica", 16), pady=20)
    label.pack(expand=True)

    # 1500ms後（1.5秒後）にsplashを閉じてmain()を起動
    splash.after(1500, lambda: [splash.destroy(), main()])
    splash.mainloop()

def open_selected_file(event=None):
    global file_path
    if file_path and os.path.exists(file_path):
        try:
            if sys.platform.startswith('darwin'):
                subprocess.call(('open', file_path))
            elif os.name == 'nt':
                os.startfile(file_path)
            elif os.name == 'posix':
                subprocess.call(('xdg-open', file_path))
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルを開けませんでした:\n{e}")

def read_excel_file(filepath):
    if filepath.endswith(".xls"):
        df = pd.read_excel(filepath, engine="xlrd")
    else:
        df = pd.read_excel(filepath, engine="openpyxl")

    return df

def detect_excel_format(filepath):
    with open(filepath, 'rb') as f:
        header = f.read(8)
    
    if header.startswith(b'\xD0\xCF\x11\xE0'):  # .xls (OLE Compound File Binary)
        return 'xls'
    elif header.startswith(b'PK'):  # .xlsx (ZIP-based)
        return 'xlsx'
    else:
        return 'unknown'
    
def attempt_repair_xls(filepath):
    try:
        df = pd.read_excel(filepath, engine="xlrd")
        save_path = filepath.replace(".xls", "_converted.xlsx")
        df.to_excel(save_path, index=False)
        messagebox.showinfo("修復成功", f"新しいExcelファイルとして保存しました:\n{save_path}")
        return save_path
    except Exception as e:
        messagebox.showerror("修復失敗", f"ファイル修復に失敗しました:\n{e}")
        return None

def open_file():
    global wb, df, file_path
    clear_columns_display()
    selected_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])

    if selected_path:
        file_path = selected_path
        filename = os.path.basename(file_path)
        file_label.config(text=f"{filename}", fg="blue", cursor="hand2")
        file_label.bind("<Button-1>", lambda e: os.startfile(file_path))
        ToolTip(file_label, "クリックでファイルを開く")

        if file_path.endswith(".xls"):
            try:
                df = pd.read_excel(file_path, engine="xlrd")

                # 列名がすべてUnnamedまたは欠損のときはExcel風の列名を生成
                if all([str(col).startswith("Unnamed") or pd.isna(col) for col in df.columns]):
                    df.columns = [get_excel_column_name(i + 1) for i in range(len(df.columns))]

                wb = None
                analyze_columns_xls(df)

            except Exception:
                actual_format = detect_excel_format(file_path)
                if actual_format == 'xlsx':
                    repaired = attempt_repair_xls(file_path)
                    if repaired:
                        wb = px.load_workbook(repaired, data_only=True)
                        analyze_columns_xlsx(wb)
                else:
                    messagebox.showerror("エラー", "xlsファイルが破損しているか、形式が異なっています。")
        else:
            try:
                wb = px.load_workbook(file_path, data_only=True)
                analyze_columns_xlsx(file_path)  # この中で列名処理を追加する必要あり
            except Exception as e:
                messagebox.showerror("エラー", f"ファイルを読み込めませんでした\n{e}")
    else:
        file_path = None
        file_label.config(text="ファイル未選択")
        columns_text.set("")
        return

def get_excel_column_name(n):
    """1-indexedでA, B, ..., Z, AA, AB, ... を返す"""
    name = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        name = "列" + chr(65 + rem) + name
    return name

def analyze_columns_xls(df):
    for i, col in enumerate(df.columns):
        count = pd.to_numeric(df[col], errors='coerrse').notna().sum()
        tk.Label(columns_frame, text=str(col), anchor="w", width=25).grid(row=i, column=0, sticky="w")
        tk.Label(columns_frame, text=f"{count}個", anchor="e", width=10).grid(row=i, column=1, sticky="e")
        
def analyze_columns_xlsx(filepath):
    clear_columns_display()

    try:
        if filepath.endswith(".xls"):
            df = pd.read_excel(filepath, engine="xlrd", header=None)
            for i, col in enumerate(df.columns):
                count = pd.to_numeric(df[col], errors='coerce').notna().sum()
                tk.Label(columns_frame, text=str(col), anchor="w", width=25).grid(row=i, column=0, sticky="w")
                tk.Label(columns_frame, text=f"{count}個", anchor="e", width=10).grid(row=i, column=1, sticky="e")
        else:
            wb = px.load_workbook(filepath, data_only=True)
            sheet = wb.active
            max_col = sheet.max_column
            max_row = sheet.max_row

            tk.Label(columns_frame, text="列名", font=("Helvetica", 10, "bold"), anchor="w", width=25).grid(row=0, column=0, sticky="w")
            tk.Label(columns_frame, text="データ個数", font=("Helvetica", 10, "bold"), anchor="e", width=10).grid(row=0, column=1, sticky="e")

            for col_idx in range(1, max_col + 1):
                part1 = sheet.cell(row=1, column=col_idx).value
                part2 = sheet.cell(row=2, column=col_idx).value

                # 両方ともNoneまたは空文字ならExcel列名に置き換え
                if (not part1 or str(part1).strip() == "") and (not part2 or str(part2).strip() == ""):
                    col_name = get_excel_column_name(col_idx)
                else:
                    # 片方でも値がある場合は結合（空白やスラッシュ調整）
                    col_name = f"{part1 or ''}/{part2 or ''}".strip(" /")

                count = 0
                for row in range(2, max_row + 1):
                    cell = sheet.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        count += 1

                row_num = col_idx
                tk.Label(columns_frame, text=col_name, anchor="w", width=25).grid(row=row_num, column=0, sticky="w")
                tk.Label(columns_frame, text=f"{count}件", anchor="e", width=10).grid(row=row_num, column=1, sticky="e")

    except Exception as e:
        messagebox.showerror("エラー", f"ファイルを読み込めませんでした\n{e}")

def clear_columns_display():
    global columns_frame

    for widget in columns_frame.winfo_children():
        widget.destroy()

def save_file():
    global wb, df
    if wb is None and df is None:
        messagebox.showwarning("エラー", "まずExcelファイル(*.xlsxまたは*.xls)を開いてくださいヽ(`Д´)ﾉ")
        return

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="名前を付けて保存"
    )

    if not save_path:
        return  # キャンセルされた場合は何もしない

    try:
        if save_path.endswith(".xls"):
            messagebox.showwarning("保存形式エラー", "xls形式には保存できません。xlsx形式で保存してください。")
            return

        if df is not None:
            df.to_excel(save_path, index=False)
        else:
            wb.save(save_path)

        messagebox.showinfo("保存完了", f"Excelファイルを保存しました:\n{save_path}")
        print(f"ファイルを保存しました: {save_path}")

    except Exception as e:
        messagebox.showerror("保存失敗", f"ファイルの保存に失敗しました:\n{e}")
    
def quit_app():
    tk.Tk().withdraw()
    res = messagebox.askokcancel('アプリ終了', 'アプリを終了しますか？')

    if res is True:
        root.quit()

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        x, y, _, _ = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 20
        y = y + self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # 枠なし
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, background="#ffffe0", relief="solid", borderwidth=1, font=("Helvetica", 9))
        label.pack()

    def hide_tip(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

def main():
    global root, file_label, columns_text, columns_frame

    # GUIウィンドウ作成
    root = tk.Tk()
    root.title("Excelデータ変換")

    # ウィンドウの幅と高さ
    window_width = 600
    window_height = 400

    # 画面サイズの取得
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # 中央座標の計算
    x = int((screen_width - window_width) / 2)
    y = int((screen_height - window_height) / 2)

    # ウィンドウの位置を設定
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")

    #全体を左右の2カラムに分割するためのフレーム
    left_frame = tk.Frame(root)
    left_frame.grid(row=0, column=0, sticky="n", padx=10, pady=10)

    right_frame = tk.Frame(root)
    right_frame.grid(row=0, column=1, sticky="n", padx=10, pady=10)

    # 左側の表示
    description = tk.Label(left_frame, text="Excelファイルを読み込み、\n単位変換とグラフ描画を行います(´・ω・｀)", font=("Helvetica", 9))
    description.pack(pady =(0, 10))

    open_button = tk.Button(left_frame, text="Excelファイルを開く", width=20, height=2, command=open_file)
    open_button.pack(pady=10)

    save_button = tk.Button(left_frame, text="保存先を指定して保存", width=20, height=2, command=save_file)
    save_button.pack(pady=10)

    quit_button = tk.Button(left_frame, text = "アプリを終了する", width=20, height=2, command=quit_app)
    quit_button.pack(pady=10)

    file_label = tk.Label(right_frame, text="ファイル未選択", fg="blue", font=("Helvetica", 10), cursor="hand2")
    file_label.pack(pady=(0, 10))
    file_label.bind("<Button-1>", open_selected_file)

    def on_enter(event):
        file_label.config(text=file_label.cget("text"))

    def on_leave(event):
        if file_path:
            filename = os.path.basename(file_path)
        else:
            filename = "ファイル未選択"
        file_label.config(text=filename)

    file_label.bind("<Enter>", on_enter)
    file_label.bind("<Leave>", on_leave)

    columns_text = tk.StringVar()
    tk.Label(right_frame, textvariable=columns_text, justify="left", wraplength=250, anchor="w").pack(pady=10)

    # 🧩 スクロール可能な columns_frame を定義
    canvas = tk.Canvas(right_frame, width=300, height=250)
    scrollbar = tk.Scrollbar(right_frame, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    global columns_frame
    columns_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=columns_frame, anchor="nw")

    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    columns_frame.bind("<Configure>", on_frame_configure)

    root.mainloop()


if __name__ == "__main__":
    show_splash_and_launch_main()